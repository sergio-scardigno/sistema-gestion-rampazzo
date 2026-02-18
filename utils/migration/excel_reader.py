"""
Lector de Excel con deteccion automatica de estructura por hoja.
"""
import openpyxl
import warnings

# Mapeos preconfigurados por hoja (basados en analisis del Excel de Rampazzo)
SHEET_MAPPINGS = {
    "CARPETAS": {
        "header_row": 2,
        "data_start": 3,
        "columns": {
            1: "id_carpeta",
            2: "nombre_completo",
            3: "cuil",
            4: "observaciones",
            5: "clave_mi_anses",
            6: "clave_fiscal",
            7: "fecha_apertura",
        },
    },
    "BASE DE DATOS": {
        "header_row": 1,
        "data_start": 3,
        "columns": {
            1: "nombre_completo",
            2: "direccion",
        },
    },
    "EXP IPS ": {
        "header_row": 1,
        "data_start": 3,
        "columns": {
            1: "id_carpeta",
            2: "nombre_completo",
            3: "cuil",
            4: "estado",
            5: "tipo_tramite",
            6: "fecha_apertura",
        },
    },
    "RTI DESFAVORABLES": {
        "header_row": 0,
        "data_start": 1,
        "columns": {
            1: "id_carpeta",
            2: "nombre_completo",
            3: "cuil",
            4: "clave_mi_anses",
            5: "estado",
            6: "tipo_responsable",
            7: "fecha_apertura",
            8: "numero_expediente",
            9: "fecha_control",
        },
    },
    "SEGUIMIENTO EXP": {
        "header_row": 3,
        "data_start": 4,
        "columns": {
            1: "id_carpeta",
            2: "nombre_completo",
            3: "cuil",
            4: "numero_expediente",
            5: "estado",
            6: "clave_mi_anses",
            7: "fecha_apertura",
            8: "fecha_control",
        },
    },
    "FALTA EDAD": {
        "header_row": 0,
        "data_start": 4,
        "columns": {
            1: "id_carpeta",
            2: "nombre_completo",
            3: "cuil",
            4: "observaciones",
        },
    },
    "videollamada": {
        "header_row": 0,
        "data_start": 1,
        "columns": {
            1: "id_carpeta",
            2: "nombre_completo",
            3: "cuil",
            4: "observaciones",
        },
    },
    "inicio virtual 2025": {
        "header_row": 0,
        "data_start": 2,
        "columns": {
            1: "id_carpeta",
            2: "nombre_completo",
            3: "cuil",
            4: "clave_mi_anses",
            5: "observaciones",
            6: "tipo_responsable",
            7: "fecha_apertura",
            9: "fecha_control",
        },
    },
    "TURNOS ANSES NUEVO": {
        "header_row": 0,
        "data_start": 1,
        "columns": {
            1: "id_carpeta",
            2: "observaciones",
        },
    },
}


def load_workbook(path: str):
    """Cargar workbook de Excel."""
    warnings.filterwarnings("ignore", category=UserWarning)
    return openpyxl.load_workbook(path, data_only=True)


def get_sheet_info(wb) -> list[dict]:
    """Retorna info basica de cada hoja."""
    result = []
    for name in wb.sheetnames:
        ws = wb[name]
        # Count non-empty rows
        data_rows = 0
        for row in range(1, min(ws.max_row + 1, 10000)):
            has_data = False
            for col in range(1, min(ws.max_column + 1, 5)):
                if ws.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if has_data:
                data_rows += 1

        # Get preview (first 3 data rows)
        preview = []
        mapping = SHEET_MAPPINGS.get(name, {})
        start = mapping.get("data_start", 2)
        for row in range(start, min(start + 3, ws.max_row + 1)):
            row_data = {}
            for col in range(1, min(ws.max_column + 1, 10)):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    row_data[f"Col{col}"] = str(val)[:80]
            if row_data:
                preview.append(row_data)

        result.append({
            "name": name,
            "rows": data_rows,
            "columns": ws.max_column,
            "preview": preview,
            "has_mapping": name in SHEET_MAPPINGS,
            "recommended": name in [
                "CARPETAS", "BASE DE DATOS", "EXP IPS ", "RTI DESFAVORABLES",
                "SEGUIMIENTO EXP", "FALTA EDAD", "TURNOS ANSES NUEVO"
            ],
        })
    return result


def read_sheet_data(wb, sheet_name: str, column_mapping: dict = None) -> list[dict]:
    """
    Leer datos de una hoja con mapeo de columnas.
    column_mapping: {col_index: field_name} o usa el predefinido.
    """
    ws = wb[sheet_name]
    mapping_config = SHEET_MAPPINGS.get(sheet_name, {})
    if column_mapping is None:
        column_mapping = mapping_config.get("columns", {})

    data_start = mapping_config.get("data_start", 2)
    records = []

    for row in range(data_start, ws.max_row + 1):
        record = {"_source_sheet": sheet_name, "_source_row": row}
        has_data = False

        for col_idx, field_name in column_mapping.items():
            val = ws.cell(row=row, column=int(col_idx)).value
            if val is not None:
                has_data = True
                record[field_name] = val

        if has_data:
            records.append(record)

    return records
